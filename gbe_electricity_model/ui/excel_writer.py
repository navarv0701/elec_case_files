"""
Excel integration for GBE Electricity Trading.
Writes real-time state, forecasts, and recommendations to Excel via openpyxl.
Optionally uses xlwings for live updates if available.
"""

from __future__ import annotations

import os
import logging
from datetime import datetime
from pathlib import Path

from state.game_state import GameState, Recommendation
from models.demand import demand_contracts
from models.production import solar_production
from config import (
    CUSTOMER_PRICE_PER_MWH, DISPOSAL_PENALTY, ELEC_SPOT_CONTRACT_SIZE,
    ELEC_FWD_CONTRACT_SIZE, DAYS_WITH_CLOSEOUT_FINE,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Writes game state and recommendations to Excel.

    Uses openpyxl to create/update an Excel workbook with multiple sheets.
    Each call to update() overwrites the data in-place so Excel shows
    current state when the file is opened or refreshed.
    """

    def __init__(self, filepath: str | None = None):
        self.filepath = filepath or str(
            Path(__file__).parent.parent / "excel" / "GBE_Dashboard.xlsx"
        )
        self._wb = None
        self._initialized = False

    def initialize(self):
        """Create the workbook with all sheet templates."""
        if not HAS_OPENPYXL:
            logger.warning("openpyxl not installed. Excel output disabled.")
            return

        wb = openpyxl.Workbook()

        # Create sheets
        sheets = ["Overview", "Producer", "Distributor", "Trader",
                   "Forecasts", "Actions", "Manual Input"]

        # Rename default sheet
        wb.active.title = sheets[0]
        for name in sheets[1:]:
            wb.create_sheet(name)

        # Style constants
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")

        # Setup each sheet with headers
        self._setup_overview(wb["Overview"], header_font_white, header_fill)
        self._setup_producer(wb["Producer"], header_font_white, header_fill)
        self._setup_distributor(wb["Distributor"], header_font_white, header_fill)
        self._setup_trader(wb["Trader"], header_font_white, header_fill)
        self._setup_forecasts(wb["Forecasts"], header_font_white, header_fill)
        self._setup_actions(wb["Actions"], header_font_white, header_fill)
        self._setup_manual(wb["Manual Input"], header_font_white, header_fill)

        # Ensure directory exists
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)

        wb.save(self.filepath)
        self._wb = wb
        self._initialized = True
        logger.info(f"Excel dashboard created: {self.filepath}")

    def update(self, state: GameState):
        """Update all sheets with current state."""
        if not self._initialized:
            self.initialize()
        if self._wb is None:
            return

        try:
            self._update_overview(self._wb["Overview"], state)
            self._update_producer(self._wb["Producer"], state)
            self._update_distributor(self._wb["Distributor"], state)
            self._update_trader(self._wb["Trader"], state)
            self._update_forecasts(self._wb["Forecasts"], state)
            self._update_actions(self._wb["Actions"], state)
            self._wb.save(self.filepath)
        except Exception as e:
            logger.error(f"Excel update failed: {e}")

    # ------------------------------------------------------------------
    # Sheet Setup
    # ------------------------------------------------------------------

    def _setup_overview(self, ws, font, fill):
        headers = ["Metric", "Value"]
        self._write_header_row(ws, headers, font, fill)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _setup_producer(self, ws, font, fill):
        # Section 1: Production Schedule
        ws["A1"] = "PRODUCTION SCHEDULE"
        ws["A1"].font = Font(bold=True, size=14)
        headers = ["Day", "Solar (hrs)", "Solar (contracts)", "Gas (contracts)",
                    "Total", "NG Cost", "Revenue Est"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = font
            cell.fill = fill
        # Widen columns
        for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col_letter].width = 18

    def _setup_distributor(self, ws, font, fill):
        ws["A1"] = "DEMAND & PROCUREMENT"
        ws["A1"].font = Font(bold=True, size=14)
        headers = ["Day", "Temp (C)", "Demand (contracts)", "Demand Low",
                    "Demand High", "Procured", "Deficit/Surplus", "Revenue"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = font
            cell.fill = fill
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col_letter].width = 18

    def _setup_trader(self, ws, font, fill):
        ws["A1"] = "TRADING & TENDERS"
        ws["A1"].font = Font(bold=True, size=14)
        headers = ["Type", "Ticker", "Action", "Qty", "Price",
                    "Expected PnL", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = font
            cell.fill = fill
        for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col_letter].width = 16

    def _setup_forecasts(self, ws, font, fill):
        ws["A1"] = "WEATHER FORECASTS"
        ws["A1"].font = Font(bold=True, size=14)
        headers = ["Day", "Sun Fcst 1", "Sun Fcst 2", "Sun Fcst 3",
                    "Temp Fcst 1", "Temp Fcst 2", "Temp Fcst 3",
                    "Sun Uncertainty", "Temp Uncertainty"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = font
            cell.fill = fill
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[col_letter].width = 15

    def _setup_actions(self, ws, font, fill):
        ws["A1"] = "RECOMMENDED ACTIONS"
        ws["A1"].font = Font(bold=True, size=14)
        headers = ["#", "Urgency", "Action", "Ticker", "Qty",
                    "Price", "Reason", "E[PnL]", "Risk"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = font
            cell.fill = fill
        ws.column_dimensions["G"].width = 50
        for col_letter in ["A", "B", "C", "D", "E", "F", "H", "I"]:
            ws.column_dimensions[col_letter].width = 12

    def _setup_manual(self, ws, font, fill):
        ws["A1"] = "MANUAL INPUT / OVERRIDES"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A3"] = "Use this sheet to enter manual overrides"
        ws["A5"] = "Sunshine Override (Day):"
        ws["A6"] = "Sunshine Override (Hours):"
        ws["A7"] = "Temperature Override (Day):"
        ws["A8"] = "Temperature Override (C):"
        ws["A10"] = "Notes:"
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    # ------------------------------------------------------------------
    # Sheet Updates
    # ------------------------------------------------------------------

    def _update_overview(self, ws, state: GameState):
        """Update the Overview sheet."""
        data = [
            ("Day", f"{state.current_day} / 5"),
            ("Tick", str(state.current_tick)),
            ("Ticks Left (Day)", str(state.ticks_remaining_in_day())),
            ("Ticks Left (Total)", str(state.ticks_remaining_total())),
            ("Status", state.case_status),
            ("", ""),
            ("ELEC-F Bid", f"${state.elec_fwd_bid:.0f}" if state.elec_fwd_bid > 0 else "-"),
            ("ELEC-F Ask", f"${state.elec_fwd_ask:.0f}" if state.elec_fwd_ask > 0 else "-"),
            ("NG Ask", f"${state.ng_ask:.2f}" if state.ng_ask > 0 else "-"),
            ("", ""),
            ("ELEC-F Position", str(state.elec_f_position)),
            ("NG Position", str(state.ng_position)),
            ("", ""),
            ("NLV", f"${state.nlv:,.0f}"),
            ("Realized PnL", f"${state.realized_pnl:,.0f}"),
            ("Unrealized PnL", f"${state.unrealized_pnl:,.0f}"),
            ("", ""),
            ("Last Updated", datetime.now().strftime("%H:%M:%S")),
        ]

        for i, (metric, value) in enumerate(data, 2):
            ws.cell(row=i, column=1, value=metric)
            ws.cell(row=i, column=2, value=value)

    def _update_producer(self, ws, state: GameState):
        """Update the Producer sheet."""
        for day in range(2, 7):
            row = day + 2  # Row 4 = Day 2, etc.
            sun = state.best_sunshine_forecast(day)
            solar = solar_production(sun) if sun is not None else 0
            gas = state.gas_production.get(day, 0)
            total = solar + gas

            ng_price = state.ng_ask if state.ng_ask > 0 else state.ng_last
            ng_cost = gas * 8 * ng_price * 100 if ng_price > 0 else 0

            # Revenue estimate
            elec_price = state.elec_fwd_last if state.elec_fwd_last > 0 else 40
            revenue = total * elec_price * ELEC_SPOT_CONTRACT_SIZE

            ws.cell(row=row, column=1, value=day)
            ws.cell(row=row, column=2, value=round(sun, 1) if sun is not None else "?")
            ws.cell(row=row, column=3, value=round(solar))
            ws.cell(row=row, column=4, value=gas)
            ws.cell(row=row, column=5, value=round(total))
            ws.cell(row=row, column=6, value=round(ng_cost))
            ws.cell(row=row, column=7, value=round(revenue))

    def _update_distributor(self, ws, state: GameState):
        """Update the Distributor sheet."""
        for day in range(2, 7):
            row = day + 2
            temp = state.best_temperature_forecast(day)
            if temp is not None:
                demand = demand_contracts(temp)
                temp_unc = state.temperature_uncertainty(day)
                demand_low = demand_contracts(temp - temp_unc)
                demand_high = demand_contracts(temp + temp_unc)
            else:
                demand = 180
                demand_low = 120
                demand_high = 240
                temp = None

            procured = state.electricity_purchased.get(day, 0)
            balance = procured - demand
            revenue = demand * CUSTOMER_PRICE_PER_MWH * ELEC_SPOT_CONTRACT_SIZE

            ws.cell(row=row, column=1, value=day)
            ws.cell(row=row, column=2, value=round(temp, 1) if temp is not None else "?")
            ws.cell(row=row, column=3, value=round(demand))
            ws.cell(row=row, column=4, value=round(demand_low))
            ws.cell(row=row, column=5, value=round(demand_high))
            ws.cell(row=row, column=6, value=round(procured))
            ws.cell(row=row, column=7, value=round(balance))
            ws.cell(row=row, column=8, value=round(revenue))

    def _update_trader(self, ws, state: GameState):
        """Update the Trader sheet."""
        row = 4
        # Pending tenders
        for tender in state.pending_tenders:
            ws.cell(row=row, column=1, value="Tender")
            ws.cell(row=row, column=2, value=tender.ticker)
            ws.cell(row=row, column=3, value=tender.action)
            ws.cell(row=row, column=4, value=tender.quantity)
            ws.cell(row=row, column=5, value=round(tender.price, 2))
            ws.cell(row=row, column=6, value="")
            ws.cell(row=row, column=7, value="Pending")
            row += 1

        # Positions needing closeout
        for ticker, pos in state.elec_day_positions.items():
            if pos != 0:
                ws.cell(row=row, column=1, value="Position")
                ws.cell(row=row, column=2, value=ticker)
                ws.cell(row=row, column=3, value="SELL" if pos > 0 else "BUY")
                ws.cell(row=row, column=4, value=abs(pos))
                ws.cell(row=row, column=5, value="")
                ws.cell(row=row, column=6, value=f"-${abs(pos) * DISPOSAL_PENALTY:,}")
                ws.cell(row=row, column=7, value="CLOSE")
                row += 1

    def _update_forecasts(self, ws, state: GameState):
        """Update the Forecasts sheet."""
        for day in range(2, 7):
            row = day + 2

            sun_forecasts = state.sunshine_forecasts.get(day, [])
            temp_forecasts = state.temperature_forecasts.get(day, [])

            ws.cell(row=row, column=1, value=day)

            # Sunshine forecasts
            for i, fc in enumerate(sun_forecasts[:3]):
                ws.cell(row=row, column=2 + i, value=round(fc.value, 1))

            # Temperature forecasts
            for i, fc in enumerate(temp_forecasts[:3]):
                ws.cell(row=row, column=5 + i, value=round(fc.value, 1))

            # Uncertainties
            ws.cell(row=row, column=8, value=round(state.sunshine_uncertainty(day), 1))
            ws.cell(row=row, column=9, value=round(state.temperature_uncertainty(day), 1))

    def _update_actions(self, ws, state: GameState):
        """Update the Actions sheet with current recommendations."""
        # Clear old data (rows 4-20)
        for row in range(4, 20):
            for col in range(1, 10):
                ws.cell(row=row, column=col, value="")

        for i, rec in enumerate(state.active_recommendations[:15], 1):
            row = i + 3
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=rec.urgency)
            ws.cell(row=row, column=3, value=rec.action)
            ws.cell(row=row, column=4, value=rec.ticker)
            ws.cell(row=row, column=5, value=rec.quantity)
            ws.cell(row=row, column=6, value=f"${rec.price:.0f}" if rec.price else "MKT")
            ws.cell(row=row, column=7, value=rec.reason)
            ws.cell(row=row, column=8, value=round(rec.expected_pnl) if rec.expected_pnl else "")
            ws.cell(row=row, column=9, value=round(rec.penalty_risk) if rec.penalty_risk else "")

    # ------------------------------------------------------------------
    # Utility
    # ------------------------------------------------------------------

    def _write_header_row(self, ws, headers, font, fill):
        """Write a header row with styling."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = font
            cell.fill = fill
